#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
  construir_snapshot.py
  --------------------------------------------------------------------------
  Lee el Excel y guarda 'snapshot.json' con:
     - el listado de partidos (fila, equipos, dia, hora)
     - los vaticinios de cada participante, ligados a la fila del partido

  IMPORTANTE: lee los nombres de equipo desde las FORMULAS (=+Mexico), porque
  openpyxl no recalcula formulas y al subir el Excel a la nube los valores
  calculados pueden venir vacios. Las formulas SIEMPRE estan, asi que de ahi
  extraemos el nombre de forma confiable.
"""
import openpyxl, datetime, json, re
from pathlib import Path

EXCEL_PATH = "2026_QUINIELA_CONSOLIDADO.xlsm"
HOJA_MARCADORES = "Marcadores_FINALES"
EXC = {"Hoja1","EDITAR (2)","Instrucciones",HOJA_MARCADORES,"Tabla 1","Banderas"}

def is_time(v): return isinstance(v, datetime.time)

def team_name(value_calc, value_formula):
    """Devuelve el nombre del equipo. Prefiere el valor calculado; si esta
    vacio, lo extrae de la formula tipo '=+Corea_del_Sur'."""
    if value_calc not in (None, "", " "):
        s = str(value_calc).strip()
        if s and not s.startswith("="):
            return s
    if isinstance(value_formula, str) and value_formula.startswith("=+"):
        return value_formula[2:].replace("_", " ").strip()
    return None

def main():
    if not Path(EXCEL_PATH).exists():
        print("No encuentro", EXCEL_PATH); return
    wb  = openpyxl.load_workbook(EXCEL_PATH, data_only=True)   # valores calculados
    wbf = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)    # formulas
    mf  = wb[HOJA_MARCADORES]
    mff = wbf[HOJA_MARCADORES]

    master = []
    for r in range(1, mf.max_row+1):
        b=mf.cell(r,2).value; c=mf.cell(r,3).value
        d = team_name(mf.cell(r,4).value, mff.cell(r,4).value)
        f = team_name(mf.cell(r,6).value, mff.cell(r,6).value)
        if is_time(b) and d and f:
            dia = c.date().isoformat() if hasattr(c,"date") else str(c)[:10]
            master.append({"row":r,"h":b.strftime("%H:%M"),"d":dia,"l":d,"v":f})

    players=[]
    for s in wb.sheetnames:
        if s in EXC: continue
        ws=wb[s]; wsf=wbf[s]
        # nombre del participante (C2), con respaldo al nombre de la hoja
        name = ws["C2"].value
        if name in (None, "", " "):
            name = s
        name = str(name).strip()
        preds=[]
        for r in range(1, ws.max_row+1):
            hf=wsf.cell(r,8).value      # formula H -> =+Marcadores_FINALES!$E$<fila>
            e=ws.cell(r,5).value; g=ws.cell(r,7).value
            if isinstance(hf,str) and "Marcadores_FINALES" in hf:
                m=re.search(r"\$E\$(\d+)", hf)
                if m and e is not None and g is not None:
                    preds.append({"mrow":int(m.group(1)),"pl":e,"pv":g})
        players.append({"n":name,"sheet":s,"preds":preds})

    Path("snapshot.json").write_text(
        json.dumps({"master":master,"players":players}, ensure_ascii=False),
        encoding="utf-8")
    print(f"snapshot.json creado: {len(master)} partidos, {len(players)} participantes.")

if __name__ == "__main__":
    main()

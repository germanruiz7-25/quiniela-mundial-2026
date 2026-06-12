#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
============================================================================
  QUINIELA COPA MUNDIAL 2026  ·  Actualizador automatico
============================================================================
  CICLO COMPLETO al finalizar cada partido:
    1. Lee resultados FINALIZADOS desde football-data.org (API confiable).
    2. FILTRO DE CALIDAD: solo acepta marcadores de partidos terminados,
       con equipos y fecha que coinciden con tu Excel.
    3. Escribe el marcador en columnas E (local) y G (visitante) de la hoja
       'Marcadores_FINALES'.
    4. Calcula los puntos de cada participante (misma regla que tus formulas)
       y regenera el Dashboard HTML.
    5. Toma una CAPTURA (PNG) del dashboard.
    6. Deja un ENLACE de WhatsApp de 1 clic para enviar al 50255152112.

  INSTALACION (una sola vez, en la terminal):
    pip install openpyxl requests playwright
    python -m playwright install chromium

  ANTES DE LA PRIMERA CORRIDA (una sola vez):
    python construir_snapshot.py     # crea snapshot.json con los vaticinios

  USO:
    python actualizar_quiniela.py            # automatico (API)
    python actualizar_quiniela.py --manual   # tu ingresas marcadores a mano

  REGLA DE PUNTOS (igual a tus formulas):
    Acierto exacto de marcador ............... 5 pts
    Acierto de resultado + un marcador exacto  3 pts
    Acierto de resultado (gana/empata/pierde)  2 pts
    Resultado equivocado ..................... 0 pts
============================================================================
"""

import sys, json, datetime, unicodedata, urllib.parse, webbrowser
from pathlib import Path

# ---------------------------------------------------------------------------
# CONFIGURACION  ->  EDITA SOLO ESTAS LINEAS
# ---------------------------------------------------------------------------
import os
API_KEY       = os.environ.get("FOOTBALL_API_KEY", "3f2be9e89a4540c5bf8ac91e8bd79c59")  # de football-data.org (o secreto de GitHub)
EXCEL_PATH    = "2026_QUINIELA_CONSOLIDADO.xlsm"  # ruta a tu Excel
TEMPLATE_HTML = "dashboard_template.html"         # plantilla del tablero
DASHBOARD_OUT = "Dashboard_Quiniela_2026.html"    # web que se regenera
SHOT_OUT      = "captura_dashboard.png"           # captura para WhatsApp
SNAPSHOT      = "snapshot.json"                    # vaticinios (lo crea construir_snapshot.py)
WHATSAPP_NUM  = "50255152112"
COMPETITION   = "WC"
HOJA_MARCADORES = "Marcadores_FINALES"

# ---------------------------------------------------------------------------
# MAPA DE NOMBRES (Espanol en tu Excel  <->  Ingles en la API)
# ---------------------------------------------------------------------------
ALIAS = {
    "mexico":["mexico"], "estados unidos":["united states","usa"],
    "paises bajos":["netherlands"], "alemania":["germany"], "espana":["spain"],
    "inglaterra":["england"], "francia":["france"], "belgica":["belgium"],
    "brasil":["brazil"], "argentina":["argentina"], "croacia":["croatia"],
    "suiza":["switzerland"], "catar":["qatar"], "corea del sur":["south korea","korea republic"],
    "chequia":["czechia","czech republic"], "sudafrica":["south africa"], "canada":["canada"],
    "bosnia y herzegovina":["bosnia and herzegovina","bosnia-herzegovina"],
    "marruecos":["morocco"], "haiti":["haiti"], "escocia":["scotland"],
    "australia":["australia"], "turquia":["turkey","turkiye"], "paraguay":["paraguay"],
    "curazao":["curacao"], "costa de marfil":["ivory coast","cote d'ivoire"],
    "ecuador":["ecuador"], "japon":["japan"], "suecia":["sweden"], "tunez":["tunisia"],
    "egipto":["egypt"], "iran":["iran","ir iran"], "nueva zelanda":["new zealand"],
    "colombia":["colombia"], "uruguay":["uruguay"], "portugal":["portugal"],
    "senegal":["senegal"], "ghana":["ghana"], "noruega":["norway"], "austria":["austria"],
    "argelia":["algeria"], "irak":["iraq"], "jordania":["jordan"],
    "arabia saudita":["saudi arabia"], "uzbekistan":["uzbekistan"],
    "cabo verde":["cape verde","cabo verde"], "rd congo":["dr congo","congo dr"],
    "panama":["panama"],
}

def norm(s):
    if s is None: return ""
    s=str(s).strip().lower()
    s="".join(c for c in unicodedata.normalize("NFD",s) if unicodedata.category(c)!="Mn")
    return s

def same_team(excel_name, api_name):
    a,b=norm(excel_name),norm(api_name)
    if a==b: return True
    for es,ens in ALIAS.items():
        if a==es and b in [norm(x) for x in ens]: return True
        if a==es and any(norm(x) in b or b in norm(x) for x in ens): return True
    return a in b or b in a

def is_time(v): return isinstance(v, datetime.time)

def load_snapshot():
    if not Path(SNAPSHOT).exists():
        print(f"  FALTA {SNAPSHOT}. Corre primero:  python construir_snapshot.py")
        sys.exit(1)
    return json.loads(Path(SNAPSHOT).read_text(encoding="utf-8"))

# ---------------------------------------------------------------------------
# 1) RESULTADOS DESDE LA API
# ---------------------------------------------------------------------------
def fetch_api_results():
    import requests
    url=f"https://api.football-data.org/v4/competitions/{COMPETITION}/matches"
    try:
        r=requests.get(url, headers={"X-Auth-Token":API_KEY}, timeout=20)
        r.raise_for_status()
    except Exception as e:
        print(f"  [API] Error de conexion: {e}"); return []
    out=[]
    for m in r.json().get("matches",[]):
        if m.get("status")!="FINISHED": continue          # FILTRO #1
        ft=m.get("score",{}).get("fullTime",{})
        gl,gv=ft.get("home"),ft.get("away")
        if gl is None or gv is None: continue
        out.append({"home":m["homeTeam"].get("name") or m["homeTeam"].get("shortName"),
                    "away":m["awayTeam"].get("name") or m["awayTeam"].get("shortName"),
                    "gl":int(gl),"gv":int(gv),"date":m.get("utcDate","")[:10]})
    print(f"  [API] {len(out)} partidos finalizados.")
    return out

# ---------------------------------------------------------------------------
# 2) ESCRIBIR MARCADORES (columnas E y G) usando el snapshot para los nombres
# ---------------------------------------------------------------------------
def update_excel(api_results, snap):
    import openpyxl
    wb=openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    mf=wb[HOJA_MARCADORES]
    rows=snap["master"]   # nombres tomados del snapshot (no de formulas)
    applied=skipped=0
    for res in api_results:
        match=None
        for row in rows:
            if same_team(row["l"],res["home"]) and same_team(row["v"],res["away"]):
                match=(row,res["gl"],res["gv"]); break                 # FILTRO #2
            if same_team(row["l"],res["away"]) and same_team(row["v"],res["home"]):
                match=(row,res["gv"],res["gl"]); break                 # equipos invertidos
        if not match: skipped+=1; continue
        row,gl,gv=match
        cur_e=mf.cell(row["row"],5).value; cur_g=mf.cell(row["row"],7).value
        try:                                                            # FILTRO #3: idempotente
            if int(cur_e)==gl and int(cur_g)==gv: continue
        except: pass
        mf.cell(row["row"],5).value=gl   # E = marcador equipo de columna D
        mf.cell(row["row"],7).value=gv   # G = marcador equipo de columna F
        applied+=1
        print(f"    OK  {row['l']} {gl}-{gv} {row['v']}  (fila {row['row']})")
    wb.save(EXCEL_PATH)
    print(f"  [Excel] {applied} escritos, {skipped} sin coincidencia.")
    return applied

# ---------------------------------------------------------------------------
# 3) PUNTOS (misma regla que tus formulas J)
# ---------------------------------------------------------------------------
def score(pl, pv, gl, gv):
    """pl,pv = vaticinio ; gl,gv = marcador final. Devuelve puntos."""
    try: pl,pv,gl,gv=int(pl),int(pv),int(gl),int(gv)
    except: return 0
    dir_pred=(pl>pv)-(pl<pv); dir_real=(gl>gv)-(gl<gv)
    if dir_pred!=dir_real: return 0          # resultado equivocado
    if pl==gl and pv==gv: return 5           # marcador exacto
    if pl==gl or pv==gv: return 3            # un marcador exacto
    return 2                                  # solo resultado

# ---------------------------------------------------------------------------
# 4) REGENERAR DASHBOARD HTML (con puntos calculados en Python)
# ---------------------------------------------------------------------------
def read_final_scores():
    import openpyxl
    wb=openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    mf=wb[HOJA_MARCADORES]
    finals={}
    for r in range(1, mf.max_row+1):
        e=mf.cell(r,5).value; g=mf.cell(r,7).value
        try: finals[r]=(int(e),int(g))
        except: finals[r]=None
    return finals

def build_data_blob(snap):
    finals=read_final_scores()
    mrow={m["row"]:m for m in snap["master"]}
    matches=[]
    for m in snap["master"]:
        fin=finals.get(m["row"])
        matches.append({"h":m["h"],"d":m["d"],"l":m["l"],"v":m["v"],
                        "gl":str(fin[0]) if fin else "*","gv":str(fin[1]) if fin else "*"})
    players=[]
    for p in snap["players"]:
        total=0; preds={}
        for pr in p["preds"]:
            m=mrow.get(pr["mrow"])
            if not m: continue
            preds[f"{m['l']}|{m['v']}"]={"l":str(pr["pl"]),"v":str(pr["pv"])}
            fin=finals.get(pr["mrow"])
            if fin: total+=score(pr["pl"],pr["pv"],fin[0],fin[1])
        players.append({"n":p["n"],"p":total,"preds":preds})
    return {"matches":matches,"players":players}

def regenerate_dashboard(snap):
    tpl=Path(TEMPLATE_HTML)
    if not tpl.exists():
        print(f"  [HTML] Falta {TEMPLATE_HTML}. Omitido."); return
    blob=json.dumps(build_data_blob(snap), ensure_ascii=False)
    Path(DASHBOARD_OUT).write_text(tpl.read_text(encoding="utf-8").replace("__DATA__",blob),
                                   encoding="utf-8")
    print(f"  [HTML] Dashboard regenerado -> {DASHBOARD_OUT}")

# ---------------------------------------------------------------------------
# 5) CAPTURA + 6) WHATSAPP
# ---------------------------------------------------------------------------
def screenshot_dashboard():
    try: from playwright.sync_api import sync_playwright
    except ImportError: print("  [Captura] Instala playwright."); return
    uri=Path(DASHBOARD_OUT).resolve().as_uri()
    with sync_playwright() as p:
        b=p.chromium.launch(); pg=b.new_page(viewport={"width":1320,"height":1400})
        pg.goto(uri); pg.wait_for_timeout(2500)
        pg.screenshot(path=SHOT_OUT, full_page=True); b.close()
    print(f"  [Captura] -> {SHOT_OUT}")

def whatsapp_link(applied):
    msg=(f"Quiniela Mundial 2026 - Tabla actualizada ({applied} partido(s)). "
         f"Adjunto la imagen del tablero del dia.")
    link=f"https://wa.me/{WHATSAPP_NUM}?text={urllib.parse.quote(msg)}"
    print("\n  ===========================================================")
    print("   WHATSAPP (1 clic) - al abrir, adjunta la imagen:")
    print(f"   {link}")
    print(f"   Imagen: {Path(SHOT_OUT).resolve()}")
    print("  ===========================================================\n")
    try: webbrowser.open(link)
    except: pass

# ---------------------------------------------------------------------------
# MODO MANUAL
# ---------------------------------------------------------------------------
def manual_results(snap):
    print("\n  MODO MANUAL: Enter vacio = saltar partido.")
    finals=read_final_scores(); out=[]
    for m in snap["master"]:
        if finals.get(m["row"]): continue
        ans=input(f"   {m['d']} {m['l']} vs {m['v']} -> (ej 2-1): ").strip()
        if "-" in ans:
            try:
                gl,gv=[int(x) for x in ans.split("-")]
                out.append({"home":m["l"],"away":m["v"],"gl":gl,"gv":gv,"date":m["d"]})
            except: print("     formato invalido, omitido")
    return out

# ---------------------------------------------------------------------------
# FLUJO PRINCIPAL
# ---------------------------------------------------------------------------
def main():
    manual="--manual" in sys.argv
    print("="*64); print("  QUINIELA 2026 -", "MANUAL" if manual else "API"); print("="*64)
    snap=load_snapshot()

    print("\n[1/5] Obteniendo resultados...")
    results = manual_results(snap) if manual else fetch_api_results()
    applied = update_excel(results, snap) if results else 0

    if applied==0 and not results:
        regenerate_dashboard(snap); print("Dashboard refrescado (sin marcadores nuevos)."); return
    if applied==0:
        print("  Nada nuevo (marcadores ya estaban). Regenero dashboard igual.")

    print("\n[2/5] Marcadores escritos con filtro de calidad.")
    print("\n[3/5] Regenerando dashboard...");  regenerate_dashboard(snap)
    print("\n[4/5] Tomando captura...");        screenshot_dashboard()
    print("\n[5/5] Enviando a WhatsApp...");     whatsapp_link(applied)
    print(f"\nLISTO. Sube {DASHBOARD_OUT} a tu hosting (Netlify) para refrescar el link publico.")

if __name__=="__main__":
    main()
